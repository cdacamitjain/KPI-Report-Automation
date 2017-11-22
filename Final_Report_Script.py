"""
# # DATE		: 12,September 2017
# # AUTHOR		: AMIT.JAIN@LNTTECHSERVICES.COM
# # DESCRIPTION	: This script is used to create Final_KPI_Report.xlsx inside ./Report Directory,
# # 			  And uses all bundle*.csv files present inside ./Referral_Data Directory.
"""
from string import whitespace

from openpyxl import load_workbook

import pandas as pd
import os.path
import sys
import csv
import time
import logging
import fnmatch

from Combine_Configs_Script import CreateCombineConfigs
from Constants import Constants
from Formats import Formats
from Formulas import Formulas
from Utils import Utils


class FinalKPIReport:
    __Const = None
    CONFIG_HEADER = [u'Test Case No.', u'Test Category', u'Test Case ID', u'Test Type', u'Configurations']
    ReadTemplateFlag = True
    DF_Template = None
    DF_Bundles_List = []
    bundleFiles_list = []

    def __init__(self):
        self.__Const = Constants()
        # logging.basicConfig(format=self.__Const.LOGGING_FORMAT, level=logging.DEBUG)
        logging.debug("constructor is called...")
        Utils.ChangeDirPath(self.__Const.REFERRAL_DATA_PATH)

        self.bundleFiles_list = list(
            reversed(fnmatch.filter(os.listdir(os.getcwd()), self.__Const.WILDCARD_BUNDLE_FILE)))
        logging.debug(self.bundleFiles_list)
        if self.bundleFiles_list != []:
            logging.debug(
                "List of files present inside " + self.__Const.REFERRAL_DATA_PATH + ":" + str(self.bundleFiles_list))
        else:
            logging.error(
                self.__Const.REFERRAL_DATA_PATH + " Directory should contains 'Bundle*.csv files'. Script is aborted...!!")

        if os.path.isdir("..\\" + self.__Const.REPORT_PATH):
            pass
        else:
            logging.error(
                self.__Const.REPORT_PATH + " Directory is need to be created to Generate Report. Script is aborted...!!")
            sys.exit(0)

        if len(self.bundleFiles_list) <= 1:
            sys.exit("Script is aborted...Cannot Generate Report for Single Bundle.")
        self.ReadBundles(self.bundleFiles_list)

    def ReadBundles(self, list_bundleFiles):
        for files in list_bundleFiles:
            if os.path.isfile(os.path.realpath(files)):
                logging.debug(files + " File is present")
                if self.ReadTemplateFlag:
                    self.DF_Template = Utils.ReadCSV(self.__Const.CONFIG_TEMPLATE_FILE)
                    self.ReadTemplateFlag = False
                self.DF_Bundles_List.append(Utils.ReadCSV(files))
            else:
                logging.error(files + " File not present")
                logging.error(
                    self.__Const.WILDCARD_BUNDLE_FILE + " File is needed to create Final Report. Script is aborted...!!")
                sys.exit(0)

    def CreateSheets(self,Max_RunCount):
        df_template = self.DF_Template
        AppsDict, AppNames = Utils.MakeApplicationDictionary(df_template)
        AppCount = 0
        Utils.ChangeDirPath(Constants.REPORT_PATH)

        writer = pd.ExcelWriter(Constants.FINAL_REPORT_FILE, engine='xlsxwriter')
        workbook = writer.book

        # # # Create Summary Tabs
        self.Create_Summary_Tab(workbook, AppNames)

        # # # # Create Module Details BUNDLE_MD Tabs
        self.Create_Summary_MD_Tab(workbook, Constants.SUMMARY_BUNDLE_MD_TAB_NAME)

        # # # # Create Module Details CONFIG_MD Tabs
        self.Create_Summary_MD_Tab(workbook, Constants.SUMMARY_CONFIG_MD_TAB_NAME)

        # # # Create Module Details Tabs
        self.Create_ModuleDetail_Tab(workbook, AppNames)

        # # Create Module Details _MD Tabs
        self.Create_ModuleDetail_MD_Tab(workbook, AppNames)

        # # # Create All Application and Application_MD Tabs
        self.Create_ApplicationMD_Tab(workbook, AppNames, AppsDict, df_template,Max_RunCount)
        workbook.close()

    def Create_Summary_MD_Tab(self, workbook, TabName):
        Row_Index = 0
        Col_Index = 0

        if TabName == Constants.SUMMARY_BUNDLE_MD_TAB_NAME:
            worksheet = workbook.add_worksheet(Constants.SUMMARY_BUNDLE_MD_TAB_NAME)
            PerformanceList = [Constants.NUMBER_OF_TESTS] + Constants.PERFORMANCE_STATUS_HEADER
            Color = Constants.COLOR_LIGHT_BLUE
        else:
            worksheet = workbook.add_worksheet(Constants.SUMMARY_CONFIG_MD_TAB_NAME)
            PerformanceList = [Constants.NUMBER_OF_TESTS] + Constants.PERFORMANCE_STATUS_RESULT_HEADER
            Color = Constants.COLOR_GOLD

        worksheet.set_column("D:AAA", Constants.CELL_WIDTH)
        worksheet.set_column("A:B", Constants.REPORT_DESCRIPTION_WIDTH)

        worksheet.merge_range(Row_Index, Col_Index, Row_Index, Col_Index + 1, Constants.CONFIGURATIONS_DETAIL_TITLE,
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_NAVY_BLUE))

        worksheet.write(Row_Index + 1, Col_Index, Constants.CONFIGURATIONS_NAME_HEADER,
                        Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))

        worksheet.write(Row_Index + 1, Col_Index + 1, Constants.CONFIGURATIONS_DESCRIPTION_HEADER,
                        Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))

        # # # Average Response Time_Table
        if TabName == Constants.SUMMARY_BUNDLE_MD_TAB_NAME:
            self.AverageResponseTime_Table(worksheet,workbook,Row_Index,Col_Index,PerformanceList)

        for i in range(Constants.CONFIGURATIONS_COUNT):
            for j in range(2):
                if j == 0:
                    worksheet.write(Row_Index + 2 + i, Col_Index + j, Constants.CONFIGURATIONS_LIST[i],
                                    Formats.Format_Cell(workbook))
                else:
                    worksheet.write(Row_Index + 2 + i, Col_Index + j, Constants.CONFIGURATIONS_DESCRIPTION_LIST[i],
                                    Formats.Format_Cell(workbook))

        worksheet.merge_range(Row_Index + 1, Col_Index + 3, Row_Index + 1, Col_Index + (2 * len(PerformanceList)) + 3,
                              Constants.SUMMARY_MD_TAB_TITLE,
                              Formats.Format_Hyperlink(workbook))

        worksheet.merge_range(Row_Index + 2, Col_Index + 3, Row_Index + 2, Col_Index + (2 * len(PerformanceList)) + 3,
                              "Improvement Count",
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_GREY))
        row = Row_Index + 12
        col = Col_Index + 3

        for i in range(len(Constants.SUMMARY_MD_TAB_HEADER)):
            if i == 0:
                worksheet.merge_range(row, col, row + 1, col, Constants.CONFIGURATIONS_NAME_HEADER,
                                      Formats.Format_First_MD_Header(workbook, Color))
            else:
                worksheet.merge_range(row, col + 1, row, col + len(PerformanceList), Constants.SUMMARY_MD_TAB_HEADER[i],
                                      Formats.Format_First_MD_Header(workbook, Color))
                for j in range(len(PerformanceList)):
                    worksheet.write(row + 1, col + 1 + j, PerformanceList[j],
                                    Formats.Format_First_MD_Header(workbook, Color))
                col += len(PerformanceList)

        row = Row_Index + 14
        col = Col_Index + 3

        for i in range(Constants.CONFIGURATIONS_COUNT + 1):
            if i < Constants.CONFIGURATIONS_COUNT:
                for j in range(len(PerformanceList) * 2 + 1):
                    if j == 0:
                        worksheet.write(row + i, col + j, Constants.CONFIGURATIONS_LIST[i],
                                        Formats.Format_Cell(workbook))
                    elif j == 1:
                        start_cell = Utils.GetColumnName(col + 1 + j)+ str(row + 1 + i)
                        end_cell = Utils.GetColumnName(col + len(PerformanceList) - 1 + j) + str(row + 1 + i)
                        formula_str = '=SUM({0}:{1})'.format(start_cell,end_cell)

                        worksheet.write(row + i, col + j, formula_str, Formats.Format_Cell(workbook))
                    elif j <= len(PerformanceList):
                        if TabName==Constants.SUMMARY_CONFIG_MD_TAB_NAME:
                            p_col = Col_Index - 1 + j + len(PerformanceList)
                        else:
                            p_col = Col_Index - 1 + j
                        formula_str = '={0}!{1}{2}'.format(Constants.MODULE_DETAILS_MD_TAB_NAME,
                                                           Utils.GetColumnName(p_col),
                                                           Constants.PERFORMANCE_ROW_INDEX_REFERENCE_TABLE + 1 + i)

                        worksheet.write(row + i, col + j, formula_str, Formats.Format_Cell(workbook))
                    else:
                        worksheet.write(row + i, col + j, "", Formats.Format_Cell(workbook))
            else:
                for j in range(len(PerformanceList) * 2 + 1):
                    if j == 0:
                        worksheet.write(row + i, col + j, "TOTAL",
                                        Formats.Format_Cell(workbook))
                    else:
                        formula_str = "=SUM({0}15:{0}18)".format(Utils.GetColumnName(col + j))
                        worksheet.write(row + i, col + j, formula_str, Formats.Format_Cell(workbook))


        #### Create Current release PIE CHART
        formula_format = '${0}${1}'
        f_row = row
        f_col = col

        start_cat_cell =formula_format.format(Utils.GetColumnName(f_col+2),f_row)
        end_cat_cell =formula_format.format(Utils.GetColumnName(f_col+len(PerformanceList)),f_row)

        f_row+=Constants.CONFIGURATIONS_COUNT + 1

        start_val_cell =formula_format.format(Utils.GetColumnName(f_col+2),f_row)
        end_val_cell =formula_format.format(Utils.GetColumnName(f_col+len(PerformanceList)),f_row)

        f_col += len(PerformanceList)

        # print start_cat_cell,end_cat_cell,start_val_cell,end_val_cell

        chartPie = workbook.add_chart(
            {'type': 'pie'})  # Configure the series. Note the use of the list syntax to define ranges:
        chartPie.add_series({
            'name': 'Pie data',
            'data_labels': {'percentage': True},
            'categories': '=' + TabName + '!{0}:{1}'.format(start_cat_cell,end_cat_cell),
            'values': '=' + TabName + '!{0}:{1}'.format(start_val_cell,end_val_cell),

        })
        chartPie.set_size({'width': 400, 'height': 150})
        # Add a title.
        chartPie.set_title({'name': Constants.CURRENT_RELEASE_HEADER})
        # Set an Excel chart style. Colors with white outline and shadow.
        chartPie.set_style(10)
        # Insert the chart into the worksheet (with an offset).
        worksheet.insert_chart('F4', chartPie)

        #### Create Previous release PIE CHART
        s_row = row
        s_col = f_col

        start_cat_cell =formula_format.format(Utils.GetColumnName(s_col+2),s_row)
        end_cat_cell =formula_format.format(Utils.GetColumnName(s_col+len(PerformanceList)),s_row)

        s_row+=Constants.CONFIGURATIONS_COUNT + 1

        start_val_cell =formula_format.format(Utils.GetColumnName(s_col+2),s_row)
        end_val_cell =formula_format.format(Utils.GetColumnName(s_col+len(PerformanceList)),s_row)

        # print start_cat_cell,end_cat_cell,start_val_cell,end_val_cell

        chartPie = workbook.add_chart(
            {'type': 'pie'})  # Configure the series. Note the use of the list syntax to define ranges:
        chartPie.add_series({
            'name': 'Pie data',
            'data_labels': {'percentage': True},
            'categories': '=' + TabName+ '!{0}:{1}'.format(start_cat_cell,end_cat_cell),
            'values': '=' + TabName + '!{0}:{1}'.format(start_val_cell,end_val_cell),

        })
        chartPie.set_size({'width': 400, 'height': 150})
        # Add a title.
        chartPie.set_title({'name': Constants.PREVIOUS_RELEASE_HEADER})
        # Set an Excel chart style. Colors with white outline and shadow.
        chartPie.set_style(10)
        # Insert the chart into the worksheet (with an offset).
        worksheet.insert_chart('K4', chartPie)

    def Create_Summary_Tab(self, workbook, AppNames):
        Row_Index = 0
        Col_Index = 0

        worksheet = workbook.add_worksheet(Constants.SUMMARY_TAB_NAME)
        worksheet.set_column("E:AAA", Constants.CELL_WIDTH)
        worksheet.set_column("A:B", Constants.REPORT_DESCRIPTION_WIDTH)
        worksheet.set_column("D:D", Constants.REPORT_DESCRIPTION_WIDTH)

        # # TEST_REPORT_DESCRIPTION_TITLE
        worksheet.merge_range(Row_Index, Col_Index, Row_Index + 3, Col_Index + 1,
                              Constants.TEST_REPORT_DESCRIPTION_TITLE,
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_NAVY_BLUE))

        for i in range(len(Constants.TEST_REPORT_DESCRIPTION_INDEX_LIST)):
            worksheet.write(Row_Index + 4 + i, Col_Index, Constants.TEST_REPORT_DESCRIPTION_INDEX_LIST[i],
                            Formats.Format_Cell(workbook))
            worksheet.write(Row_Index + 4 + i, Col_Index + 1, "",
                            Formats.Format_Cell(workbook))

        # # Module Detail Table
        col = Col_Index
        for i in range(len(Constants.SUMMARY_TAB_HEADER)):
            if i == 0:
                worksheet.merge_range(Row_Index + len(Constants.TEST_REPORT_DESCRIPTION_INDEX_LIST), col + 3,
                                      Row_Index + len(Constants.TEST_REPORT_DESCRIPTION_INDEX_LIST) + 1, Col_Index + 3,
                                      Constants.SUMMARY_TAB_HEADER[i],
                                      Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))
            else:
                worksheet.merge_range(Row_Index + len(Constants.TEST_REPORT_DESCRIPTION_INDEX_LIST),
                                      col,
                                      Row_Index + len(Constants.TEST_REPORT_DESCRIPTION_INDEX_LIST),
                                      col + len(Constants.SUMMARY_TAB_SUB_HEADER) - 1,
                                      Constants.SUMMARY_TAB_HEADER[i],
                                      Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))
                for j in range(len(Constants.SUMMARY_TAB_SUB_HEADER)):
                    worksheet.write(Row_Index + len(Constants.TEST_REPORT_DESCRIPTION_INDEX_LIST) + 1, col + j,
                                    Constants.SUMMARY_TAB_SUB_HEADER[j],
                                    Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))

            col += len(Constants.SUMMARY_TAB_SUB_HEADER)

        Write_Col = Col_Index + len(Constants.SUMMARY_TAB_SUB_HEADER) - 1
        Write_Row = Row_Index + len(Constants.TEST_REPORT_DESCRIPTION_INDEX_LIST) + 2
        col = Write_Col
        row = Write_Row

        for i in range(len(AppNames) + 1):
            Reference_SheetName = Constants.MODULE_DETAILS_TAB_NAME
            if i < len(AppNames):
                for j in range(2 * len(Constants.SUMMARY_TAB_SUB_HEADER) + 1):
                    if j == 0:
                        worksheet.write(Write_Row + i, Write_Col + j, AppNames[i], Formats.Format_Cell(workbook))
                    elif j <= len(Constants.SUMMARY_TAB_SUB_HEADER):
                        # print Write_Row+i,Write_Col+j
                        # =SUM(Module_Details!H3:H5)
                        formula_str = '=SUM({0}!{1}{2}:{1}{3})'.format(Reference_SheetName, Utils.GetColumnName(3 + j),
                                                                       col, col + 2)
                        worksheet.write(Write_Row + i, Write_Col + j, formula_str, Formats.Format_Cell(workbook))
                        # print formula_str
                    else:
                        worksheet.write(Write_Row + i, Write_Col + j, "", Formats.Format_Cell(workbook))
                col += 3
            else:
                for j in range(2 * len(Constants.SUMMARY_TAB_SUB_HEADER) + 1):
                    if j == 0:
                        worksheet.write(Write_Row + i, Write_Col + j, "TOTAL", Formats.Format_Cell(workbook))
                    else:
                        formula_str = "=SUM({}{}:{}{})".format(Utils.GetColumnName(Write_Col + j),
                                                               Write_Row + i - len(AppNames) + 1,
                                                               Utils.GetColumnName(Write_Col + j), Write_Row + i)
                        # print formula_str
                        worksheet.write(Write_Row + i, Write_Col + j, formula_str, Formats.Format_Cell(workbook))
                        row = Write_Row + i

        #### Create Current release PIE CHART

        chartPie = workbook.add_chart({'type': 'pie'})
        # Configure the series. Note the use of the list syntax to define ranges:
        chartPie.add_series({
            'name': 'Pie data',
            'data_labels': {'percentage': True},
            'categories': '=' + Constants.SUMMARY_TAB_NAME + '!$F$14:$H$14',
            'values': '=' + Constants.SUMMARY_TAB_NAME + '!$F$' + str(row + 1) + ':$H$' + str(row + 1),

        })
        chartPie.set_size({'width': 420, 'height': 220})
        # Add a title.
        chartPie.set_title({'name': Constants.CURRENT_RELEASE_HEADER})
        # Set an Excel chart style. Colors with white outline and shadow.
        chartPie.set_style(10)
        # Insert the chart into the worksheet (with an offset).
        worksheet.insert_chart('F1', chartPie, {'x_offset': 1, 'y_offset': 10})

        #### Create Previous release PIE CHART

        chartPie = workbook.add_chart({'type': 'pie'})
        # Configure the series. Note the use of the list syntax to define ranges:
        chartPie.add_series({
            'name': 'Pie data',
            'data_labels': {'percentage': True},
            'categories': '=' + Constants.SUMMARY_TAB_NAME + '!$J$14:$L$14',
            'values': '=' + Constants.SUMMARY_TAB_NAME + '!$J$' + str(row + 1) + ':$L$' + str(row + 1),

        })
        chartPie.set_size({'width': 420, 'height': 220})
        # Add a title.
        chartPie.set_title({'name': Constants.PREVIOUS_RELEASE_HEADER})

        # Set an Excel chart style. Colors with white outline and shadow.
        chartPie.set_style(10)

        # Insert the chart into the worksheet (with an offset).
        worksheet.insert_chart('J1', chartPie, {'x_offset': 1, 'y_offset': 10})

    def Create_ModuleDetail_MD_Tab(self, workbook, AppNames):
        # print AppNames
        Row_Index = 0
        Col_Index = 0
        BundleName_list = Utils.GetBundlesList(self.bundleFiles_list)
        CurrentBundleName = BundleName_list[0]

        worksheet = workbook.add_worksheet(Constants.MODULE_DETAILS_MD_TAB_NAME)
        worksheet.set_column("A:AAA", Constants.CELL_WIDTH)

        Header_Col_Index = Col_Index
        # Application Name Header
        worksheet.merge_range(Row_Index, Header_Col_Index, Row_Index + 1, Header_Col_Index,
                              Constants.APPLICATION_NAME_HEADER,
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_GREY))

        # Configuration Header
        Header_Col_Index += 1
        worksheet.merge_range(Row_Index, Header_Col_Index, Row_Index + 1, Header_Col_Index,
                              Constants.CONFIGURATIONS_HEADER,
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_GREY))

        # Bundle to Bundle Performance Result Comparison for Current Bundle Header
        Header_Col_Index += 1
        worksheet.merge_range(Row_Index, Header_Col_Index, Row_Index,
                              Header_Col_Index + len(Constants.PERFORMANCE_STATUS_HEADER) - 1,
                              Constants.BUNDLE_TO_BUNDLE_COMPARISON_HEADER.format(CurrentBundleName),
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))

        for i in range(len(Constants.PERFORMANCE_STATUS_HEADER)):
            worksheet.write(Row_Index + 1, Header_Col_Index + i, Constants.PERFORMANCE_STATUS_HEADER[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))

        Header_Col_Index += len(Constants.PERFORMANCE_STATUS_HEADER)

        # Config to Config Comparison Performance Result for Current Bundle Header
        worksheet.merge_range(Row_Index, Header_Col_Index, Row_Index,
                              Header_Col_Index + len(Constants.PERFORMANCE_STATUS_RESULT_HEADER) - 1,
                              Constants.CONFIG_TO_CONFIG_COMPARISON_HEADER.format(CurrentBundleName),
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_GOLD))

        for i in range(len(Constants.PERFORMANCE_STATUS_RESULT_HEADER)):
            worksheet.write(Row_Index + 1, Header_Col_Index + i, Constants.PERFORMANCE_STATUS_RESULT_HEADER[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_GOLD))

        Header_Col_Index += len(Constants.PERFORMANCE_STATUS_RESULT_HEADER)

        # Bundle Performance header
        worksheet.merge_range(Row_Index, Header_Col_Index, Row_Index, Header_Col_Index + 1,
                              Constants.BUNDLES_PERFORMANCE_TITLE,
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_ORANGE))
        for i in range(len(Constants.BUNDLES_PERFORMANCE_HEADER)):
            worksheet.write(Row_Index + 1, Header_Col_Index + i, Constants.BUNDLES_PERFORMANCE_HEADER[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_ORANGE))
            Header_Col_Index += i


        # Average Response Time
        self.BundleAverageResTimeHeader_Col(worksheet, workbook, Row_Index, Header_Col_Index + 1,
                                            (Constants.AVERAGE_RESULTS_HEADER + BundleName_list))

        Write_Row_Index = Row_Index + 2
        Write_Col_Index = Col_Index
        row = Write_Row_Index
        col = Write_Col_Index
        AppCount = 0

        p_row = row + 1
        p_col = 2 + len(Constants.PERFORMANCE_STATUS_HEADER) \
                + len(Constants.PERFORMANCE_STATUS_RESULT_HEADER)\
                + len(Constants.BUNDLES_PERFORMANCE_HEADER)\
                +len(Constants.AVERAGE_RESULTS_HEADER)
        bundle_count = len(BundleName_list)

        Count = 0
        for appName in AppNames:
            Updated_AppName = Utils.SupportedSheetName(appName, 11)
            AppCount += 1
            SheetName_MD = Utils.SheetNameFormator(AppCount, Updated_AppName, "_MD")
            formula_str = Constants.HYPERLINK_FORMULA_STRING_FORMAT.format(SheetName_MD + "!A1", appName)
            worksheet.merge_range(row, col, row + Constants.CONFIGURATIONS_COUNT - 1, col, formula_str,
                                  Formats.Format_Hyperlink(workbook))
            for i in range(len(Constants.CONFIGURATIONS_LIST)):
                # CONFIG
                worksheet.write(row + i, col + 1, Constants.CONFIGURATIONS_LIST[i], Formats.Format_Cell(workbook))

                for j in range(len(Constants.PERFORMANCE_STATUS_HEADER) + len(
                        Constants.PERFORMANCE_STATUS_RESULT_HEADER) + len(Constants.BUNDLES_PERFORMANCE_HEADER) + len(
                            Constants.AVERAGE_RESULTS_HEADER + BundleName_list)):
                    if j < len(Constants.PERFORMANCE_STATUS_HEADER + Constants.PERFORMANCE_STATUS_RESULT_HEADER):
                        formula_str = '={}!{}{}'.format(SheetName_MD, Utils.GetColumnName(col + j + 2),
                                                        (Constants.PERFORMANCE_ROW_INDEX_REFERENCE_TABLE + 1 + i))
                    elif j == len(Constants.PERFORMANCE_STATUS_HEADER + Constants.PERFORMANCE_STATUS_RESULT_HEADER):
                        formula_str = Formulas.Formula_PerformanceComparedToAllReleases(p_row+Count,p_col,bundle_count,Constants.CONFIGURATIONS_COUNT,False)[0]
                    elif j == len(Constants.PERFORMANCE_STATUS_HEADER + Constants.PERFORMANCE_STATUS_RESULT_HEADER) + 1:
                        formula_str = Formulas.Formula_PerformanceComparedToLastRelease(p_row+Count,p_col)
                        Count+=1
                    else:
                        formula_str = '={}!{}{}'.format(SheetName_MD, Utils.GetColumnName(col + j),
                                                        (Constants.PERFORMANCE_ROW_INDEX_REFERENCE_TABLE + 1 + i))

                    worksheet.write(row + i, col + j + 2, formula_str, Formats.Format_Cell(workbook))

            row += Constants.CONFIGURATIONS_COUNT

        # # Performance Count Reference Table
        Col_Index_Refer = 0
        Row_Index_Refer = Constants.PERFORMANCE_ROW_INDEX_REFERENCE_TABLE
        Read_Row_Index = 3

        for i in range(Constants.CONFIGURATIONS_COUNT):
            worksheet.write(Row_Index_Refer + i, Col_Index_Refer, Constants.CONFIGURATIONS_LIST[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_GREY))
        Col_Index_Refer += 1

        # BUNDLE_TO_BUNDLE_COMPARISON_HEADER
        worksheet.merge_range(Row_Index_Refer - 2, Col_Index_Refer, Row_Index_Refer - 2,
                              Col_Index_Refer + len(Constants.PERFORMANCE_STATUS_HEADER) - 1,
                              Constants.BUNDLE_TO_BUNDLE_COMPARISON_HEADER.format(Utils.GetBundlesList(self.bundleFiles_list)[0]),
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))
        for i in range(len(Constants.PERFORMANCE_STATUS_HEADER)):
            worksheet.write(Row_Index_Refer - 1, Col_Index_Refer + i, Constants.PERFORMANCE_STATUS_HEADER[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))

        Col_Index_Refer += len(Constants.PERFORMANCE_STATUS_HEADER)

        # CONFIG_TO_CONFIG_COMPARISON_HEADER
        worksheet.merge_range(Row_Index_Refer - 2, Col_Index_Refer, Row_Index_Refer - 2,
                              Col_Index_Refer + len(Constants.PERFORMANCE_STATUS_HEADER) - 2,
                              Constants.CONFIG_TO_CONFIG_COMPARISON_HEADER.format(Utils.GetBundlesList(self.bundleFiles_list)[0]),
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_GOLD))

        for i in range(len(Constants.PERFORMANCE_STATUS_RESULT_HEADER)):
            worksheet.write(Row_Index_Refer - 1, Col_Index_Refer + i, Constants.PERFORMANCE_STATUS_RESULT_HEADER[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_GOLD))

        Col_Index_Refer += len(Constants.PERFORMANCE_STATUS_RESULT_HEADER)

        # Average Response Time for All Bundles
        worksheet.merge_range(Row_Index_Refer - 2, Col_Index_Refer, Row_Index_Refer - 2,
                              Col_Index_Refer + len(BundleName_list)+len(Constants.AVERAGE_RESULTS_HEADER) - 1,
                              Constants.BUNDLES_AVERAGE_RESPONSE_TIME,
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_GREEN))

        bundles_results_list = Constants.AVERAGE_RESULTS_HEADER+BundleName_list
        for i in range(len(bundles_results_list)):
            worksheet.write(Row_Index_Refer - 1, Col_Index_Refer + i,
                            bundles_results_list[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_GREEN))

        data_len = len(Constants.PERFORMANCE_STATUS_HEADER) + len(Constants.PERFORMANCE_STATUS_RESULT_HEADER) + len(
                bundles_results_list)
        Col_Index_Refer = 1
        # print data_len
        for i in range(Constants.CONFIGURATIONS_COUNT):
            for j in range(data_len):

                if j < data_len-len(bundles_results_list):
                    formula_str = Formulas.Formula_BundleWiseAverage(Utils.GetColumnName(Col_Index_Refer + j + 1),
                                                                     Read_Row_Index + i, len(AppNames),
                                                                     Constants.CONFIGURATIONS_COUNT, Constants.SUM_KEY)
                    worksheet.write(Row_Index_Refer + i, Col_Index_Refer + j, formula_str,
                                    Formats.Format_Cell(workbook))
                elif j >= data_len-len(bundles_results_list) and j < data_len-len(bundles_results_list)+2:
                    formula_str = Formulas.Formula_BundleWiseAverage(Utils.GetColumnName(Col_Index_Refer + j + 3),
                                                                     Read_Row_Index + i, len(AppNames),
                                                                     Constants.CONFIGURATIONS_COUNT, Constants.SUM_KEY)
                    worksheet.write(Row_Index_Refer + i, Col_Index_Refer + j, formula_str,
                                Formats.Format_Cell(workbook))
                else:

                    formula_str = Formulas.Formula_BundleWiseAverage(Utils.GetColumnName(Col_Index_Refer + j + 3),
                                                                     Read_Row_Index + i, len(AppNames),
                                                                     Constants.CONFIGURATIONS_COUNT,
                                                                     Constants.AVERAGE_KEY)
                    worksheet.write(Row_Index_Refer + i, Col_Index_Refer + j, formula_str,
                                    Formats.Format_Cell(workbook))



    def Create_ApplicationMD_Tab(self, workbook, AppNames, AppsDict, df_template,Max_RunCount):
        AppCount = 0
        for appName in AppNames:
            # for appName, tc_count in AppsDict.items():

            tc_count = AppsDict.get(appName)

            AppCount += 1
            Updated_AppName = Utils.SupportedSheetName(appName, 11)
            df_sorted = df_template.loc[[appName], :]

            df_app_list = self.RearrangeDataFrame(appName,Max_RunCount)
            run_count = Max_RunCount
            # print df_app_list
            SheetName_MD = Utils.SheetNameFormator(AppCount, Updated_AppName, "_MD")
            SheetName_APP = Utils.SheetNameFormator(AppCount, Updated_AppName)

            # # APP TAB CREATION
            self.Create_TAB(workbook, tc_count, df_sorted, df_app_list, run_count,
                            Constants.TEST_CASE_DESCRIPTION_COUNT,
                            SheetName_APP,
                            Constants.TEST_CASE_DESCRIPTION_LIST,
                            Constants.APP_TAB_HEADER,
                            Constants.TC_DESCRIPTION_WIDTH)
            # # MD TAB CREATION
            self.Create_TAB(workbook, tc_count, df_sorted, df_app_list, run_count,
                            Constants.CONFIGURATIONS_COUNT,
                            SheetName_MD,
                            Constants.CONFIGURATIONS_LIST,
                            Constants.MD_TAB_HEADER)
            # break

    def Create_ModuleDetail_Tab(self, workbook, AppNames):
        col_index = 0
        row_index = 0
        ref_col_index = col_index
        ref_row_index = row_index

        worksheet = workbook.add_worksheet(Constants.MODULE_DETAILS_TAB_NAME)
        worksheet.set_column("A:AAA", Constants.CELL_WIDTH)

        len_ModuleDetailTabHeader = len(Constants.MODULE_DETAILS_TAB_HEADER)

        for i in range(len_ModuleDetailTabHeader - 1):
            worksheet.merge_range(row_index, ref_col_index, row_index + 1, ref_col_index,
                                  Constants.MODULE_DETAILS_TAB_HEADER[i],
                                  Formats.Format_First_MD_Header(workbook, Constants.COLOR_NAVY_BLUE))
            ref_col_index += 1
        worksheet.merge_range(row_index, ref_col_index, row_index, ref_col_index + 2,
                              Constants.MODULE_DETAILS_TAB_HEADER[-1],
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_NAVY_BLUE))

        result_index_list = Constants.TEST_LEVEL_INDEX_LIST[1:-1]
        for i in range(len(result_index_list)):
            worksheet.write(row_index + 1, ref_col_index + i, result_index_list[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_NAVY_BLUE))

        write_row_index = row_index + 2
        for i in range(len(AppNames)):
            SheetName_APP = Utils.SheetNameFormator(i + 1, Utils.SupportedSheetName(AppNames[i], 11))
            col_name = 9
            for j in range(len(Constants.TEST_LEVEL_HEADER_LIST)):
                worksheet.write(write_row_index + j, col_index + 1, Constants.TEST_LEVEL_HEADER_LIST[j],
                                Formats.Format_Cell(workbook))
                worksheet.write(write_row_index + j, col_index + 2, "", Formats.Format_Cell(workbook))
                worksheet.write(write_row_index + j, col_index + 3, "", Formats.Format_Cell(workbook))

                # TOTAL
                formula_str = "={}!{}7".format(SheetName_APP, Utils.GetColumnName(col_name + j))
                worksheet.write(write_row_index + j, col_index + 4, formula_str, Formats.Format_Cell(workbook))

                # PASS
                formula_str = "={}!{}4".format(SheetName_APP, Utils.GetColumnName(col_name + j))
                worksheet.write(write_row_index + j, col_index + 5, formula_str, Formats.Format_Cell(workbook))

                # FAIL
                formula_str = "={}!{}5".format(SheetName_APP, Utils.GetColumnName(col_name + j))
                worksheet.write(write_row_index + j, col_index + 6, formula_str, Formats.Format_Cell(workbook))

                # OPEN
                formula_str = "={}!{}6".format(SheetName_APP, Utils.GetColumnName(col_name + j))
                worksheet.write(write_row_index + j, col_index + 7, formula_str, Formats.Format_Cell(workbook))

            # HYPERLINK FOR APPNAME
            formula_str = Constants.HYPERLINK_FORMULA_STRING_FORMAT.format(SheetName_APP + "!A1", AppNames[i])
            worksheet.merge_range(write_row_index, col_index, write_row_index + 2, col_index, formula_str,
                                  Formats.Format_Hyperlink(workbook))
            write_row_index += 3

    def RearrangeDataFrame(self, appName,Max_RunCount):
        first_time_flag = True
        df_FinalFrames_list = []
        current_run_count = 0
        for df in self.DF_Bundles_List:
            df_sorted = df.loc[[appName], :]
            if first_time_flag:
                current_run_count = Utils.CalculateRunCount(df_sorted)
            dataframe_list = []
            for i in range(len(df_sorted)):
                row = list(df_sorted.iloc[i, :])[Constants.BUNDLE_FIXED_HEADER_COUNT:]
                # print appName,len(row)
                split_list_count = Constants.ISSUE_HEADER_COUNT + Max_RunCount
                temp = [row[i:i + split_list_count] for i in xrange(0, len(row), split_list_count)]

                if first_time_flag:
                    dataframe_list.append(pd.DataFrame(zip(*temp)).transpose())
                else:
                    # pd.DataFrame(zip(*temp)).transpose()
                    dataframe_list.append(pd.DataFrame(zip(*temp)).transpose().iloc[:, [-3]])
                    # dataframe_list.append(pd.DataFrame(zip(*temp)).transpose())

            first_time_flag = False
            finalFrame = pd.concat(dataframe_list, ignore_index=False)

            df_FinalFrames_list.append(finalFrame)

        return df_FinalFrames_list

    def Create_TAB(self, workbook, tc_count, df_sorted, df_app_list, run_count,
                   ConfigPerTC_Count,
                   SheetName,
                   Test_Decription_list,
                   header,
                   DescWidth=Constants.CELL_WIDTH):

        worksheet = workbook.add_worksheet(SheetName)
        worksheet.set_column("A:AAA", Constants.CELL_WIDTH)
        worksheet.set_column("E:E", DescWidth)
        worksheet.set_column("G:G", DescWidth)

        Col_Index = 0
        Row_Index = 0
        Row_Index_Write = 2

        self.Header_MD(worksheet, workbook, header, Row_Index, Col_Index)  # r = 0; c = 0;
        self.TestCaseNumber_Col(worksheet, workbook, tc_count, ConfigPerTC_Count, Row_Index_Write,
                                Col_Index)  # r = 2; c = 0
        Col_Index += 1

        self.TestCategory_Col(worksheet, workbook, df_sorted.iloc[:, 1], ConfigPerTC_Count,
                              Row_Index_Write, Col_Index)  # r = 2;  c = 1
        Col_Index += 1

        self.TestCaseID_Col(worksheet, workbook, df_sorted.iloc[:, 0], ConfigPerTC_Count, Row_Index_Write,
                            Col_Index)  # r = 2;c = 2
        Col_Index += 1

        self.TestType_Col(worksheet, workbook, df_sorted.iloc[:, 2], ConfigPerTC_Count, Row_Index_Write,
                          Col_Index)  # r = 2;c = 3
        Col_Index += 1

        self.Config_Col(worksheet, workbook, tc_count, Test_Decription_list, Row_Index_Write, Col_Index)  # r = 2;c = 4
        Col_Index += 1

        # print "df_app_list",len(df_app_list)
        # print i,Utils.GetColumnName(i)

        bundles_List = Utils.GetBundlesList(self.bundleFiles_list)
        BundleCount = len(bundles_List)
        currentBundle_Name = bundles_List[0]
        df_currentBundle = df_app_list[0]

        if "_MD" in SheetName:

            Col_Index = self.PerformanceHeader_Col(worksheet, workbook, Row_Index, Col_Index, currentBundle_Name,
                                                   tc_count, run_count, BundleCount, df_currentBundle)

            self.MeasurementsAndAvgResponseTime(worksheet, workbook, Row_Index_Write, Col_Index, df_app_list, run_count)
            Col_Index = self.MeasurementHeader_Col(worksheet, workbook, Row_Index, Col_Index, currentBundle_Name,
                                                   run_count)

            Col_Index = self.BundleAverageResTimeHeader_Col(worksheet, workbook, Row_Index, Col_Index, bundles_List)

            self.PerformanceCountReferenceTable(worksheet, workbook, Row_Index_Write, Col_Index, tc_count, bundles_List,run_count)

        else:
            result_list = self.TCResult_list(Row_Index_Write, SheetName, tc_count)
            # print len(result_list)
            start_cell = (Utils.GetColumnName(Col_Index)+str(Row_Index_Write + 1))
            end_cell = (Utils.GetColumnName(Col_Index)+str(Row_Index_Write + 1+len(result_list)))
            cell_range = "{0}:{1}".format(start_cell,end_cell)
            # print cell_range
            for i in range(len(result_list)):
                worksheet.write(Row_Index_Write + i, Col_Index, result_list[i], Formats.Format_Cell(workbook))
                worksheet.write(Row_Index_Write + i, Col_Index + 1, " ", Formats.Format_Cell(workbook))

                worksheet.conditional_format(cell_range,
                                             {'type': 'cell',
                                              'criteria': 'equal to',
                                              'value': 'FAIL',
                                              'format': Formats.Format_First_MD_Header(workbook,
                                                                                       Constants.COLOR_LIGHT_ORANGE)})
                worksheet.conditional_format(cell_range,
                                         {'type': 'cell',
                                          'criteria': 'equal to',
                                          'value': 'PASS',
                                          'format': Formats.Format_First_MD_Header(workbook,
                                                                                   Constants.COLOR_LIGHT_GREEN)})
                worksheet.conditional_format(cell_range,
                                             {'type': 'cell',
                                              'criteria': 'equal to',
                                              'value': 'OPEN',
                                              'format': Formats.Format_First_MD_Header(workbook,
                                                                                       Constants.COLOR_NAVY_BLUE)})

            TestType_Dict = Utils.AppsNameWithTestCount(list(df_sorted.iloc[:, 1]))
            self.ResultReferenceTable(worksheet, workbook, Row_Index_Write, Col_Index, TestType_Dict)

    def ResultReferenceTable(self, worksheet, workbook, Row_Index_Write, Col_Index, TestType_Dict):
        Read_Col_Index = Col_Index
        Read_Row_Index = Row_Index_Write
        write_col_index = Col_Index + 3
        for i in range(len(Constants.TEST_LEVEL_INDEX_LIST)):
            worksheet.write(Row_Index_Write + i, write_col_index, Constants.TEST_LEVEL_INDEX_LIST[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_GREY))
        write_col_index += 1
        for i in range(len(Constants.TEST_LEVEL_HEADER_LIST)):
            worksheet.write(Row_Index_Write, write_col_index + i, Constants.TEST_LEVEL_HEADER_LIST[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_GREY))

        for i in range(len(Constants.TEST_LEVEL_HEADER_LIST)):
            Read_Col_Name = Utils.GetColumnName(Read_Col_Index)
            # print Read_Row_Index
            TC_count = 0
            result = Constants.TEST_LEVEL_INDEX_LIST[1:-1]
            for j in range(len(result)):
                status = result[j]
                row = Row_Index_Write + 1 + j
                col = write_col_index + i
                if Constants.TEST_LEVEL_HEADER_LIST[i] in TestType_Dict.viewkeys():
                    # print Constants.TEST_LEVEL_HEADER_LIST[i]
                    TC_count = TestType_Dict[Constants.TEST_LEVEL_HEADER_LIST[i]]

                    RowCount = TC_count * Constants.TEST_CASE_DESCRIPTION_COUNT
                    startRow = Read_Col_Name + str(Read_Row_Index + 1)
                    endRow = Read_Col_Name + str(Read_Row_Index + RowCount)
                    # print startRow,endRow
                    formula_str = '=COUNTIF({0}:{1}, "{2}")'.format(startRow, endRow, status)
                    # print row,col
                    worksheet.write(row, col, formula_str, Formats.Format_Cell(workbook))
                    # print formula_str
                else:
                    TC_count = 0
                    # print TC_count
                    # print row, col
                    worksheet.write(row, col, TC_count, Formats.Format_Cell(workbook))

            formula_str = '=SUM({0}4:{0}6)'.format(Utils.GetColumnName(write_col_index + i))
            worksheet.write(row + 1, write_col_index + i, formula_str, Formats.Format_Cell(workbook))
            # break

    def PerformanceCountReferenceTable(self, worksheet, workbook, Row_Index_Write, Col_Index, tc_count, bundles_List,run_count):
        Col_Index_Refer = 0
        Row_Index_Refer = Constants.PERFORMANCE_ROW_INDEX_REFERENCE_TABLE
        d_col = Col_Index - 12
        g_col = Col_Index - 11
        c_col = Col_Index - 3
        BundleCount = len(bundles_List)

        for i in range(Constants.CONFIGURATIONS_COUNT):
            worksheet.write(Row_Index_Refer + i, Col_Index_Refer, Constants.CONFIGURATIONS_LIST[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_GREY))
        Col_Index_Refer += 1

        worksheet.merge_range(Row_Index_Refer - 2, Col_Index_Refer, Row_Index_Refer - 1, Col_Index_Refer,
                              Constants.TOTAL_TEST_CASES,
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_GREY))

        for i in range(Constants.CONFIGURATIONS_COUNT):
            worksheet.write(Row_Index_Refer + i, Col_Index_Refer, tc_count, Formats.Format_Cell(workbook))

        Col_Index_Refer += 1

        worksheet.merge_range(Row_Index_Refer - 2, Col_Index_Refer, Row_Index_Refer - 2,
                              Col_Index_Refer + len(Constants.PERFORMANCE_STATUS_HEADER) - 1,
                              Constants.BUNDLE_TO_BUNDLE_COMPARISON_HEADER.format(
                                  Utils.GetBundlesList(self.bundleFiles_list)[0]),
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))

        e_col = Col_Index_Refer + len(Constants.PERFORMANCE_STATUS_HEADER)

        worksheet.merge_range(Row_Index_Refer - 2, e_col, Row_Index_Refer - 2,
                              e_col + len(Constants.PERFORMANCE_STATUS_RESULT_HEADER) - 1,
                              Constants.CONFIG_TO_CONFIG_COMPARISON_HEADER.format(
                                  Utils.GetBundlesList(self.bundleFiles_list)[0]),
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_GOLD))

        for i in range(len(Constants.PERFORMANCE_STATUS_RESULT_HEADER)):
            worksheet.write(Row_Index_Refer - 1, e_col + i, Constants.PERFORMANCE_STATUS_RESULT_HEADER[i] + " Count",
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_GOLD))

        last_Col_index = 0
        c2c_col = Constants.MD_TAB_HEADER_COUNT+1

        for j in range(len(Constants.PERFORMANCE_STATUS_RESULT_HEADER)):
            for i in range(Constants.CONFIGURATIONS_COUNT):
                formula_str = Formulas.Formula_PerformanceWiseCount(Row_Index_Write + 1 + i, Utils.GetColumnName(c2c_col),
                                                                    tc_count,
                                                                    Constants.CONFIGURATIONS_COUNT,
                                                                    Constants.PERFORMANCE_STATUS_RESULT_HEADER_DICT.get(
                                                                        Constants.PERFORMANCE_STATUS_RESULT_HEADER[j]))

                worksheet.write(Row_Index_Refer + i, e_col + j, formula_str, Formats.Format_Cell(workbook))
                last_Col_index = e_col + j

        worksheet.write(Constants.NA_ROW_INDEX_REFERENCE-1, 0, "NA", Formats.Format_Cell(workbook))

        e_col = Col_Index_Refer + len(Constants.PERFORMANCE_STATUS_HEADER) + len(
            Constants.PERFORMANCE_STATUS_RESULT_HEADER)

        worksheet.merge_range(Row_Index_Refer - 2, e_col, Row_Index_Refer - 2,
                              e_col + len(self.bundleFiles_list) + len(Constants.AVERAGE_RESULTS_HEADER) - 1,
                              Constants.BUNDLES_AVERAGE_RESPONSE_TIME,
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_GREEN))

        last_Col_index = 0

        b2b_col = Constants.MD_TAB_HEADER_COUNT

        for j in range(len(Constants.PERFORMANCE_STATUS_HEADER)):
            for i in range(Constants.CONFIGURATIONS_COUNT):
                formula_str = Formulas.Formula_PerformanceWiseCount(Row_Index_Write + 1 + i, Utils.GetColumnName(b2b_col),
                                                                    tc_count,
                                                                    Constants.CONFIGURATIONS_COUNT,
                                                                    Constants.PERFORMANCE_STATUS_HEADER_DICT.get(
                                                                        Constants.PERFORMANCE_STATUS_HEADER[j]))

                worksheet.write(Row_Index_Refer + i, Col_Index_Refer + j, formula_str, Formats.Format_Cell(workbook))
                last_Col_index = Col_Index_Refer + j

        for i in range(len(Constants.PERFORMANCE_STATUS_HEADER)):
            worksheet.write(Row_Index_Refer - 1, Col_Index_Refer + i, Constants.PERFORMANCE_STATUS_HEADER[i] + " Count",
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))

        Col_Index_Refer += len(Constants.PERFORMANCE_STATUS_HEADER)

        for i in range(len(Constants.AVERAGE_RESULTS_HEADER)):
            worksheet.write(Row_Index_Refer - 1, e_col + i, Constants.AVERAGE_RESULTS_HEADER[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_GREEN))

        a_col = Constants.PERFORMANCE_HEADER_COUNT + Constants.MD_TAB_HEADER_COUNT + run_count
        for i in range(Constants.CONFIGURATIONS_COUNT):
            formula_str = Formulas.Formula_TestCasesComparedForAverageResults(Row_Index_Write + 1 + i, a_col,
                                                                              BundleCount,
                                                                              Constants.CONFIGURATIONS_COUNT, tc_count)
            worksheet.write(Row_Index_Refer + i, e_col, formula_str, Formats.Format_Cell(workbook))
            # print formula_str
            formula_str = "={}-{}".format(tc_count, Utils.GetColumnName(e_col) + str(Row_Index_Refer + 1 + i))
            worksheet.write(Row_Index_Refer + i, e_col + 1, formula_str, Formats.Format_Cell(workbook))

        e_col = e_col + len(Constants.AVERAGE_RESULTS_HEADER)
        for i in range(BundleCount):
            worksheet.write(Row_Index_Refer - 1, e_col + i, bundles_List[i],
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_GREEN))

        c_col = Constants.PERFORMANCE_HEADER_COUNT + Constants.MD_TAB_HEADER_COUNT + run_count
        for i in range(Constants.CONFIGURATIONS_COUNT):
            for j in range(BundleCount):
                formula_str = Formulas.Formula_Average(Row_Index_Write + 1 + i, c_col, c_col + j, BundleCount,
                                                       Constants.CONFIGURATIONS_COUNT, tc_count)
                worksheet.write(Row_Index_Refer + i, e_col + j, formula_str, Formats.Format_Cell(workbook))

    def TCResult_list(self, Row_Index_Write, SheetName, tc_count):
        SheetName = SheetName + "_MD"
        result_list = []
        for i in range(tc_count):
            result_list += Formulas.Formula_TestCaseResult(Row_Index_Write+1, SheetName)
            Row_Index_Write += Constants.CONFIGURATIONS_COUNT
        return result_list

    def MeasurementsAndAvgResponseTime(self, worksheet, workbook, Row_Index_Write, Col_Index, df_app_list, run_count):
        first = True
        c = 0
        all_data_list = []
        for df in df_app_list:
            list_df = df.iloc[:, :run_count].fillna("NA").values.tolist()
            if first:
                first = False
                for i in range(len(list_df)):
                    for j in range(len(list_df[i])):
                        worksheet.write(Row_Index_Write + i, Col_Index + j, list_df[i][j],
                                        Formats.Format_Cell(workbook))
                df = df.iloc[:, [run_count]].fillna("NA")

            list_df = df.iloc[:, :].fillna("NA").values.tolist()
            all_data_list.append(list_df)

            for i in range(len(list_df)):
                for j in range(len(list_df[i])):
                    worksheet.write(Row_Index_Write + i, Col_Index + run_count + c, list_df[i][j],
                                    Formats.Format_Cell(workbook))
            c += 1

    def AppFreezeAndFirstMeasurement_list(self, df_CurrentBundle):
        return df_CurrentBundle.iloc[:, -2:].fillna("NA").values.tolist()

    def PerformanceComparedBetweenAllConfiguration_list(self, Row_Index_Write, Col_Index, RunCount, TC_Count):
        Row_Index_Write += 1
        str_list = []
        for i in range(TC_Count):
            str_list += Formulas.Formula_PerformanceComparedBetweenAllConfiguration(
                Utils.GetColumnName(Col_Index + Constants.PERFORMANCE_HEADER_COUNT + RunCount),
                Row_Index_Write,
                Constants.CONFIGURATIONS_COUNT)
            Row_Index_Write += Constants.CONFIGURATIONS_COUNT
        # print str_list
        return str_list

    def performanceCompareToAllReleases_list(self, Row_index, Col_index, TC_Count, RunCount, BundleCount):
        PCR_List = []
        Row_index += 1
        Col_index = Col_index + Constants.PERFORMANCE_HEADER_COUNT + RunCount

        for i in range(TC_Count):
            PCR_List += Formulas.Formula_PerformanceComparedToAllReleases(Row_index, Col_index, BundleCount,
                                                                          Constants.CONFIGURATIONS_COUNT)
            Row_index += Constants.CONFIGURATIONS_COUNT
        return PCR_List

    def BundleAverageResTimeHeader_Col(self, worksheet, workbook, Row_Index, Col_Index, bundles_List):
        worksheet.merge_range(Row_Index, Col_Index, Row_Index, Col_Index + len(bundles_List) - 1,
                              Constants.BUNDLES_AVERAGE_RESPONSE_TIME,
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_GREEN))
        for bundle in bundles_List:
            worksheet.write(Row_Index + 1, Col_Index, bundle,
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_GREEN))
            Col_Index += 1
        return Col_Index

    def MeasurementHeader_Col(self, worksheet, workbook, Row_Index, Col_Index, currentBundle_Name, run_count):
        worksheet.merge_range(Row_Index, Col_Index, Row_Index, Col_Index + run_count - 1, currentBundle_Name,
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_ORANGE))
        for i in range(run_count):
            worksheet.write(Row_Index + 1, Col_Index, Constants.MEASUREMENT_STRING_FORMAT % (i + 1),
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_ORANGE))
            Col_Index += 1
        return Col_Index

    def PerformanceHeader_Col(self, worksheet, workbook, Row_Index, Col_Index, currentBundle_Name, tc_count, RunCount,
                              BundleCount, DF_CurrentBundle):
        Reference_Index = Col_Index
        # print "Reference_Index",Reference_Index
        Row_Index_Write = Row_Index + 2
        ###### PerformanceComparedToAllReleases

        PCR_list = self.performanceCompareToAllReleases_list(Row_Index_Write, Col_Index, tc_count, RunCount,
                                                             BundleCount)
        # print PCR_list

        ######## PerformanceComparedBetweenAllConfiguration

        PCC_list = self.PerformanceComparedBetweenAllConfiguration_list(Row_Index_Write, Col_Index, RunCount, tc_count)

        ######## AppFreezeAndFirstMeasurement

        AFFM_list = self.AppFreezeAndFirstMeasurement_list(DF_CurrentBundle)

        dummyList = zip(PCR_list, PCC_list)
        for i in range(len(dummyList)):
            for j in range(len(dummyList[i])):
                worksheet.write(Row_Index_Write + i, Reference_Index + j, dummyList[i][j],
                                Formats.Format_Cell(workbook))

        Reference_Index += 2

        for i in range(len(AFFM_list)):
            for j in range(len(AFFM_list[i])+1):
                if j < 2:
                    worksheet.write(Row_Index_Write + i, Reference_Index + j, AFFM_list[i][j],
                                    Formats.Format_Cell(workbook))
                else:
                    e_col = len(Constants.PERFORMANCE_HEADER_LIST) + RunCount + Col_Index
                    formula_str = Formulas.Formula_ImprovementPercentageComparedToCurrentRelease(Row_Index_Write+1+i,e_col)
                    worksheet.write(Row_Index_Write + i, Reference_Index + j,formula_str ,
                                Formats.Format_Percentage(workbook))


        worksheet.merge_range(Row_Index, Col_Index, Row_Index, Col_Index + Constants.PERFORMANCE_HEADER_COUNT - 1,
                              currentBundle_Name,
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))
        for performanceTitle in Constants.PERFORMANCE_HEADER_LIST:
            worksheet.write(Row_Index + 1, Col_Index, performanceTitle,
                            Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_BLUE))
            Col_Index += 1

        return Col_Index

    def Header_MD(self, worksheet, workbook, header_list, r, c):
        for v in range(len(header_list)):
            worksheet.merge_range(0, c, 1, c, header_list[v],
                                  Formats.Format_First_MD_Header(workbook, Constants.COLOR_GREY))
            c += 1

    def TestCaseNumber_Col(self, worksheet, workbook, tc_count, count, r, c):
        for v in range(tc_count * count):
            worksheet.write(r, 0, v + 1, Formats.Format_Cell(workbook))
            r += 1

    def TestCategory_Col(self, worksheet, workbook, category_list, ConfigPerTC_Count, r, c):
        #     0             1                 2                   3
        # "Test Case ID" "Test Case Level" "Test Case Category" "Test Case Description"
        # r = 2;  c = 1
        for k, v in Utils.AppsNameWithTestCount(category_list).items():
            # print k, v
            worksheet.merge_range(r, c, r + (v * ConfigPerTC_Count) - 1, c, k, Formats.Format_Cell(workbook))
            r = r + (v * ConfigPerTC_Count)

    def TestCaseID_Col(self, worksheet, workbook, TestcaseId_list, ConfigPerTC_Count, r, c):
        # df_sorted.iloc[:, 0]
        # r = 2;c = 2
        for v in TestcaseId_list:
            # print v
            worksheet.merge_range(r, c, r + (1 * ConfigPerTC_Count) - 1, c, v, Formats.Format_Cell(workbook))
            r = r + (1 * ConfigPerTC_Count)

    def TestType_Col(self, worksheet, workbook, testType_list, ConfigPerTC_Count, r, c):

        for v in testType_list:
            # print v
            worksheet.merge_range(r, c, r + ConfigPerTC_Count - 1, c, v, Formats.Format_Cell(workbook))
            r = r + ConfigPerTC_Count

    def Config_Col(self, worksheet, workbook, tc_count, Config_list, r, c):

        for v in range(tc_count):
            for config in Config_list:
                worksheet.write(r, c, config, Formats.Format_Cell(workbook))
                r += 1

    def list_to_merge_cells(self, worksheet, targetList, TargetColumn, TestCaseId_List=[]):
        if TestCaseId_List:
            set_list = set(TestCaseId_List[1:])
            for item in set_list:
                first_index = TestCaseId_List.index(item)
                last_index = len(TestCaseId_List) - 1 - TestCaseId_List[::-1].index(item)
                worksheet.merge_range(TargetColumn + str(first_index + 2) + ':' + TargetColumn + str(last_index + 2),
                                      targetList[first_index])
        else:
            set_list = set(targetList[1:])
            for item in set_list:
                first_index = targetList.index(item)
                last_index = len(targetList) - 1 - targetList[::-1].index(item)
                worksheet.merge_range(TargetColumn + str(first_index + 2) + ':' + TargetColumn + str(last_index + 2),
                                      targetList[first_index])
                # logging.debug( first_index, last_index)

    def CreateApplicationMDTab(self):
        logging.debug("Creating Application MD tabs for Final KPI Report")
        # self.__Utils.ChangeDirPath(self.REPORT_PATH)
        filePath = os.path.abspath(self.__Const.FINAL_REOPRT_FILE)
        logging.debug(filePath)
        writer = pd.ExcelWriter(filePath, engine='xlsxwriter')
        df_ConfigTemp = self.ReadConfigTemplate()
        appName_List = []
        TestCategory_List = ['']
        TestCaseId_List = ['']
        TestType_List = ['']
        Config_List = ['']
        TestCaseNo_List = ['']

        for i in range(len(df_ConfigTemp.index)):
            if df_ConfigTemp['Application Name'][i] == 'CardManager':
                for j in range(len(self.CONFIG_HEADER) - 1):
                    TestType_List.append(df_ConfigTemp['Test Case Category'][i])
                    TestCaseId_List.append(df_ConfigTemp['Test Case ID'][i])
                    TestCategory_List.append(df_ConfigTemp['Test Case Level'][i])

        for i in range(len(TestCaseId_List) - 1):
            Config_List.extend(self.__Const.CONFIGURATIONS_LIST)

        for i in range(len(Config_List)):
            TestCaseNo_List.append(i + 1)

        storeDataSet = (zip(TestCaseNo_List, TestCategory_List, TestCaseId_List, TestType_List, Config_List))
        df = pd.DataFrame(data=storeDataSet, columns=self.CONFIG_HEADER)
        pd.formats.format.header_style = None
        tempPackageName_value = "CardManager_MD"
        df.to_excel(writer, tempPackageName_value, startrow=0, startcol=0, index=False)
        worksheet = writer.sheets[tempPackageName_value]

        ASCII_Value_A = 65
        for i in range(len(self.CONFIG_HEADER)):
            worksheet.merge_range(chr(ASCII_Value_A + i) + '1:' + chr(ASCII_Value_A + i) + '2',
                                  self.CONFIG_HEADER[i])

        self.list_to_merge_cells(worksheet, TestCategory_List, 'B')
        self.list_to_merge_cells(worksheet, TestCaseId_List, 'C')
        self.list_to_merge_cells(worksheet, TestType_List, 'D', TestCaseId_List)

        workbook = writer.book
        warp_text_format = workbook.add_format({'text_wrap': True})
        warp_text_format.set_center_across()
        warp_text_format.set_bold()
        warp_text_format.set_align("center")
        warp_text_format.set_align("vcenter")
        warp_text_format.set_border(2)
        # warp_text_format.set_bg_color("#C0C0C0")
        worksheet.set_column("A:E", 20, warp_text_format)

    def AverageResponseTime_Table(self, worksheet,workbook, Row_Index, Col_Index,PerformanceList):
        Row_Index += 21
        Col_Index += 3
        bundle_list = Utils.GetBundlesList(self.bundleFiles_list)
        Bundles_Average_Response_Time_List = [Constants.CONFIGURATIONS_NAME_HEADER] + Constants.AVERAGE_RESULTS_HEADER + bundle_list
        BundleCount = len(Bundles_Average_Response_Time_List)

        worksheet.merge_range(Row_Index , Col_Index , Row_Index , Col_Index + (2 * len(PerformanceList)),
                              "Average Response Time",
                              Formats.Format_First_MD_Header(workbook, Constants.COLOR_GREY))

        # # # AVERAGE TABLE
        Row_Index += 40

        for i in range(BundleCount):
            worksheet.write(Row_Index - 1, Col_Index+i ,Bundles_Average_Response_Time_List[i],
                                        Formats.Format_First_MD_Header(workbook, Constants.COLOR_LIGHT_GREEN))

        col = len(Constants.PERFORMANCE_STATUS_HEADER)+len(Constants.PERFORMANCE_STATUS_RESULT_HEADER)
        for i in range(Constants.CONFIGURATIONS_COUNT):
            for j in range(BundleCount):
                if j==0:
                    worksheet.write(Row_Index + i, Col_Index+j , Constants.CONFIGURATIONS_LIST[i],
                                        Formats.Format_Cell(workbook))
                else:
                    cell_name = Utils.GetColumnName(len(Constants.PERFORMANCE_STATUS_RESULT_HEADER)+len(Constants.PERFORMANCE_STATUS_HEADER)+j)
                    row_index = Constants.PERFORMANCE_ROW_INDEX_REFERENCE_TABLE + 1 + i
                    formula_str = "={0}!{1}{2}".format(Constants.MODULE_DETAILS_MD_TAB_NAME,cell_name,row_index)
                    worksheet.write(Row_Index + i, Col_Index+j , formula_str,
                                        Formats.Format_Cell(workbook))

        # # # CREATE AVERAGE RESPONSE TIME BAR CHART
        for i in range(Constants.CONFIGURATIONS_COUNT):
            cell_format = "${}${}"
            first_cell = cell_format.format(Utils.GetColumnName(Col_Index),str(Row_Index+1+i))
            start_cat_cell = cell_format.format(Utils.GetColumnName(Col_Index+3),str(Row_Index))
            end_cat_cell   = cell_format.format(Utils.GetColumnName(Col_Index+2+len(bundle_list)),str(Row_Index))

            start_val_cell = cell_format.format(Utils.GetColumnName(Col_Index+3),str(Row_Index+1+i))
            end_val_cell   = cell_format.format(Utils.GetColumnName(Col_Index+2+len(bundle_list)),str(Row_Index+1+i))

            # print first_cell,start_cat_cell,end_cat_cell,start_val_cell,end_val_cell

            bar_chart = workbook.add_chart({'type': 'column'})
            bar_chart.add_series({
                'name'      : '={}!{}'.format(Constants.SUMMARY_BUNDLE_MD_TAB_NAME,first_cell),
                'categories': '={}!{}:{}'.format(Constants.SUMMARY_BUNDLE_MD_TAB_NAME,start_cat_cell,end_cat_cell),
                'values'    : '={}!{}:{}'.format(Constants.SUMMARY_BUNDLE_MD_TAB_NAME,start_val_cell,end_val_cell),
            })
            bar_chart.set_style(11)

            bar_chart.set_title ({'name': 'Bundle to Bundle Comparison for {}'.format(Constants.CONFIGURATIONS_LIST[i])})
            bar_chart.set_x_axis({'name': 'Executed Bundles'})
            bar_chart.set_y_axis({'name': 'Average Time in (mSecond)'})
            bar_chart.set_size({'width': 500, 'height': 275})

            # 'F26' 'K26'
            # 'F44' 'K44'
            if i == 0:
                worksheet.insert_chart('F26', bar_chart)
            if i == 1:
                worksheet.insert_chart('K26', bar_chart)
            if i == 2:
                worksheet.insert_chart('F44', bar_chart)
            if i == 3:
                worksheet.insert_chart('K44', bar_chart)


# ________________________________________________________________________________________________
def main():
    obj = CreateCombineConfigs()
    obj.CheckConfigTemplate()
    obj.CreateCombineConfigsCSV()
    obj1 = FinalKPIReport()
    # obj1.CreateApplicationMDTab()
    time.sleep(0.25)
    obj1.CreateSheets(obj.getMaxRunCount())


# ________________________________________________________________________________________________
if __name__ == "__main__":
    main()
